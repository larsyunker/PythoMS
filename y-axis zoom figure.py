# -*- coding: utf-8 -*-
"""
Chromatogram y-axis zoomer v003
new in 003:
    - code is entirely reworked for flexible input and use
    - allows for xlsx file input
    - changed zoom ramp to be linear in magnification
    - removed right and top borders for figures
    - allows changing of font size, font, line width, and axis width
    - checks for errors in the excel file (invalid values, blank cells, etc.)

still to add:
    update to new classes, etc.
    - fix fading (previous lines aren't completely gone)
    - option for having old lines still partially visible?
    - allow definition of figure size and dpi?
"""

"""

Data should be supplied in an *.xlsx file. 
Multiple sets of data can be supplied by using multiple sheets in the excel file
The first column of each sheet should have the same unit across all sheets (e.g. time (min))"""
xlsx = 'data.xlsx' # filename of excel file containing the data to plot

# list of desired intervals for zoom transitions (in seconds)
# e.x. [2.0,5.0] will produce a 2 second first zoom transition and a 5 second second zoom transition
zoomdurs = [2.0,2.0,2.0,2.0]

# y axis values for each stop point
yvalues = [0.08, 0.0040, 0.000008,0.0000008,0.00000008]  

# desired frames per second
# (most movies are 30 frames per second)
fps = 30

# specify font size
fs = 16

# specify x-axis bounds
xbound = [0,10]

# specify linewidth
lw = 1.5

# axis line width
axwidth = 1.0

# x axis label
xlab = 'time (min)'

# y axis label
ylab = 'concentration (M)'

# show magnification? (0 = no; 1 = yes)
magni = 1

"""
 General arrangement for excel files:
     All data should be pre-processed
     Column A: time data
     Column B-on: species intensity values
     Row 1: species titles (e.g. Time(min), aldehyde, ketone, etc.)
     Row 2: desired colour for each species (either as a specified letter or a RGB set (see below for more information)
     Row 3: the transition ramp where the species should disappear (1 will disappear during the first ramp, 2 the second, etc.)
            If you want a species to always be present, ensure that it has an integer number greater than the number of ramps 
            (ex. if there are 3 ramps, a value of 4 will always be visible)
            Ensure that the time column has an integer
 
 colour definitions:
    A) supply a letter combination (ex. 'b' for blue or 'db' for dark blue)
    B) or supply a R,G,B value in '(###,###,###) format [each R/G/B value should be an integer in 0-255]
       note the ' before the brackets (otherwise excel tries to interpret these as coordinates)
 b: blue
 r: red
 g: olive green
 p: purple
 o: orange
 a: aqua
 bl: black

 d_: dark colour (ex. 'db' = dark blue)
 l_: light colour (ex. 'lb' = light blue)
 
"""

# ----------------------------------------------------------
# -------------------FUNCTION DEFINITIONS-------------------
# ----------------------------------------------------------

def secondsToStr(t):
    rediv = lambda ll,b : list(divmod(ll[0],b)) + ll[1:]
    return "%d:%02d:%02d.%02d" % tuple(reduce(rediv,[[t*1000,],1000,60,60]))

def colours(c):
    colourdict = {'b': (79,129,189),
    'r': (192,80,77),
    'g': (155,187,89),
    'p': (128,100,162),
    'a': (75,172,198),
    'o': (247,150,70),
    'lb': (149,179,215),
    'lr': (217,150,148),
    'lg': (195,214,155),
    'lp': (179,162,199),
    'la': (147,205,221),
    'lo': (250,192,144),
    'db': (54,96,146),
    'dr': (99,37,35),
    'dg': (79,98,40),
    'dp': (64,49,82),
    'da': (33,89,104),
    'do': (152,72,7),
    'bl': (0,0,0)}

    if type(c) == str:
        if colourdict.has_key(c) == False:
            sys.exit('The supplied colour "%s" is not valid.\nPlease look at the available colours and enter a valid colour call key.' %(c))
        c = colourdict[c]    
    if type(c) != str:
        if len(c) != 3:
            sys.exit('The supplied colour tuple "%s" is too long or short.\nThe RGB colour tuple should have three values (in the range 0-255) separated by commas within round brackets.\ne.g. "(79, 129, 189)" is R = 79, G = 129, B = 189' %(str(c)))
        for i in range(3):
            if type(c[i]) != int:
                sys.exit('One of the values in your colour tuple is not an integer ("%s").\nPlease correct the value.' %(str(c[i])))
            if c[i] > 255 or c[i] < 0:
                sys.exit('One of the values in your colour tuple is outside of the valid RGB range of 0-255 ("%s").\nPlease check and correct the value' %(str(c[i])))
    out = []
    out.append(round(c[0]/255.,3))
    out.append(round(c[1]/255.,3))
    out.append(round(c[2]/255.,3))
    
    return tuple(out)

def linramp(valstart,valend,dur):
    """
    Function for generating a linear ramp of values
    valstart: value at the start of the ramp
    valend: value at the end of the ramp
    dur: number of steps (duration)
    """
    out = []
    for i in range(int(dur)):
        out.append( ((float(valend-valstart))/(float(dur)))*(i) + valstart )
    return out

def mag(initial,final):
    """
    calculate magnification for a value
    """
    return float(initial)/float(final)

def linmag(vali,magstart,magend,dur):
    """
    funciton for generating a ramp of values that is linear in magnification
    vali: initial value (globally)
    magstart: magnification at the start of the ramp
    magend: magnification at the end of the ramp
    dur: number of steps (duration)
    """
    out = []
    for i in range(dur):
        out.append(float(vali)/((magend-magstart)/dur*i + magstart))
    return out

# pull and save title, color, and alpha for each species
def titlecolor(filename):
    lines = open(filename,'r')
    for i, line in enumerate(lines):
        if i == 0:
            names = line.split(',')
            names[-1] = names[-1][0:-1]
        elif i == 1:
            namecolor = line.split(',')
            namecolor[-1] = namecolor[-1][0:-1]
            for ii in range(len(namecolor)):
                if len(namecolor[ii]) > 4:
                    namecolor[ii] = namecolor[ii].split(';')
                    for iii in range(len(namecolor[ii])):
                        namecolor[ii][iii] = int(namecolor[ii][iii])
        elif i == 2:
            alphaoff = line.split(',')
            alphaoff[-1] = alphaoff[-1][0:-1]
            for i, val in enumerate(alphaoff):
                alphaoff[i] = int(alphaoff[i])
        elif i > 2:
            break
    lines.close()
    return names,namecolor,alphaoff

def renderplot(ymax,timeslot,frame):
    """
    function for rendering plots
    """
    pl.clf()
    font = {'fontname':'Arial'} #font parameters for axis/text labels
    tickfont = pl.matplotlib.font_manager.FontProperties(family='Arial',size=fs) # font parameters for axis ticks
    ax = fig.add_subplot(111)
    for ind,sheet in enumerate(data): # for each sheet
        for ind1,species in enumerate(data[sheet]): # for each species
            if data[sheet][species][1] >= 1: # if alpha trigger indicates fully visible
                if species != xname[sheet]: # if species is not x axis
                    pl.plot(data[sheet][xname[sheet]][2],data[sheet][species][2],color=colours(data[sheet][species][0]),alpha=1,linewidth=lw)
            if data[sheet][species][1] == 0: # if alpha trigger indicates fade
                if species != xname[sheet]:
                    pl.plot(data[sheet][xname[sheet]][2],data[sheet][species][2],color=colours(data[sheet][species][0]),alpha=alpharamp[frame-1],linewidth=lw)
    
    for axis in ["top","bottom","left","right"]:
        ax.spines[axis].set_linewidth(axwidth)
    pl.tick_params(axis='y', length=axwidth*3, width=axwidth, direction='out',right='off')
    pl.tick_params(axis='x', length=axwidth*3, width=axwidth, direction='out',top='off')
    for label in ax.get_xticklabels():
            label.set_fontproperties(tickfont)
    for label in ax.get_yticklabels():
            label.set_fontproperties(tickfont)
    pl.xlabel(xlab, fontsize=fs, **font)                
    pl.ylabel(ylab, fontsize=fs, **font)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    
    # x axis parameters
    pl.xlim(xbound)
    
    # y axis general parameters 
    pl.ylim(0,ymax)
    
    # display magnification
    if magni == 1:
        magn = mag(yvalues[0],ymax)
        if magn <= 2.:
            pl.text(xbound[1]*0.95,ymax*0.95,str(round(magn,1))+u'×',color='k', horizontalalignment='right', verticalalignment='center', alpha = 1.0,fontsize=fs,**font)
        if magn > 2.:
            pl.text(xbound[1]*0.95,ymax*0.95,str(round(magn,0))[:-2]+u'×',color='k', horizontalalignment='right', verticalalignment='center', alpha = 1.0,fontsize=fs,**font)
        
    #padding adjustment
    pl.subplots_adjust(left = (0.17), right = 0.97, bottom = (0.12), top = 0.95)
    #pl.tight_layout(pad=0.5)
    
    #save figure
    if frame == 0:
        pl.savefig(os.getcwd()+r'\imgs\hold'+str(timeslot)+'.png',figsize = (19.2,10.8),dpi=200)
    if frame >= 1:
        frame = 1000+frame
        pl.savefig(os.getcwd()+r'\imgs\zoom'+str(timeslot)+r'\frame'+str(frame)[1:4]+'.png',figsize = (19.2,10.8),dpi=200)

def alpha(index):
    """
    function for taking an index and converting it to the corresponding column used by excel
    works up to column ZY (column # 701)
    """
    import string
    alphabet = list(string.ascii_lowercase)
    out = ''
    n = (index+1)/26
    if n != 0:
        out += alphabet[n-1]
        out += alphabet[(index+1)%26-1]
    else:
        out = alphabet[index]
    return out


# ----------------------------------------------------------
# -------------------PROGRAM BEGINS-------------------------
# ----------------------------------------------------------

import os, sys, time, gc, imp
import pylab as pl
from ast import literal_eval
try:
    import openpyxl as op
    import lxml
except ImportError:
    sys.exit('\nOne of either openpyxl or lxml does not appear to be installed.\nPlease install these packages.')

try:
    wb = op.load_workbook(xlsx) #try loading specified excel workbook
except:
    sys.exit('\nThe specified excel file could not be loaded.\nPlease check that the name of the file ("%s") and the current working directory are correct.' %xlsx)

data = {}
xname = {}
for sheet in wb.get_sheet_names():
    rd = wb.get_sheet_by_name(sheet)
    data[sheet] = {}
    for ind,col in enumerate(rd.columns):
        if ind == 0: # if column is x axis, define key for x-axis column
            xname[sheet] = str(col[0].value)
        data[sheet]['%s' %str(col[0].value)] = [] # create key for row
        for i in range(3): # create colour, disappear, and values lists for key
            data[sheet]['%s' %str(col[0].value)].append([])
        for ind1,row in enumerate(rd.rows):
            if ind1 == 1: # colour definition row
                try: # try pulling tuple for RGB
                    data[sheet]['%s' %str(col[0].value)][0] = literal_eval(rd.cell(row = ind1+1,column = ind+1).value)
                except SyntaxError: # if not tuple, pull string
                    data[sheet]['%s' %str(col[0].value)][0] = str(rd.cell(row = ind1+1,column = (ind+1)).value)
                except ValueError: # if cell is empty, assume black
                    data[sheet]['%s' %str(col[0].value)][0] = 'k'
            elif ind1 == 2: # disappear row
                try:
                    data[sheet]['%s' %str(col[0].value)][1] = int(rd.cell(row = ind1+1,column = (ind+1)).value)
                except ValueError:
                    sys.exit('\nA non-integer value was encountered in Row 3, Column %s of %s (%s).\nEnsure that all values in row 3 are integers.' %(alpha(ind).upper(),sheet,rd.cell(row = ind1+1,column = (ind+1)).value))
            elif ind1 > 2: # data
                try:
                    data[sheet]['%s' %str(col[0].value)][2].append(float(rd.cell(row = (ind1+1),column = (ind+1)).value))
                except ValueError:
                    sys.exit('\nA invalid entry was encountered in Row %i, Column %s of %s (%s).\nNumbers are expected in any rows 4 or above.' %(ind1+1,alpha(ind).upper(),sheet,rd.cell(row = (ind1+1),column = (ind+1)).value))

# check for /img directory and create if missing
if os.path.isdir('imgs') == False:
    os.makedirs('imgs')
# check for and create folders for each zoom period
for i,val in enumerate(zoomdurs):
    if os.path.isdir(os.getcwd()+"/imgs/zoom"+str(i+1)) == False:
        os.makedirs(os.getcwd()+"/imgs/zoom"+str(i+1))

## check for appropriate combinations of zoomtimes and zoomvalues
#if len(zoomtimes) != (len(zoomvalues)-1):
#    if len(zoomtimes) > (len(zoomvalues)-1):
#        sys.exit('\nThere are too many transition times specified (zoomtimes).\nEither add a zoom value or remove a transition time.')
#    if len(zoomtimes) < (len(zoomvalues)-1):
#        sys.exit('\nThere are too many zoom values specified (zoomvalues).\nEither remove a zoom value or add a transition time.')

sys.stdout.write('Start time: %s\n' %(time.strftime('%I:%M:%S %p')))
time_zero = time.time()

# create figure and add subplot
#close()
fig = pl.figure(figsize = (9.6,5.4),dpi= 100)

for ind,val in enumerate(zoomdurs):
    current_time = time.time()
    sys.stdout.write('\rProcessing hold #%i Elapsed time: %ss' %(ind+1,secondsToStr(current_time-time_zero)))
    sys.stdout.flush()
    
    renderplot(yvalues[ind],ind+1,0) # render hold frame
    sys.stdout.write(' DONE\n')
    sys.stdout.flush()
    
    for ind2,sheet in enumerate(data): # subtract alpha triggers
        for ind3, species in enumerate(data[sheet]):
            data[sheet][species][1] -= 1
    
    # define ramps
    # syntax: initial(global),magstart, magend, duration
    zoomramp = linmag(yvalues[0],mag(yvalues[0],yvalues[ind]),mag(yvalues[0],yvalues[ind+1]),int(zoomdurs[ind]*fps))
    # start,end,duration
    alpharamp = linramp(1,0,int(zoomdurs[ind]*fps))
    # render ramp frames
    for frame,zoomval in enumerate(zoomramp):
        current_time = time.time()
        sys.stdout.write('\rProcessing zoom #%i (frame #%i/%i %.1f%%) Elapsed time: %ss' %(ind+1,frame+1,fps*val,float((frame+1)/(fps*val)*100),secondsToStr(current_time-time_zero)))
        sys.stdout.flush()
        renderplot(zoomval,ind+1,frame+1)
    sys.stdout.write(' DONE\n')
    sys.stdout.flush()
       
current_time = time.time()
sys.stdout.write('\rProcessing final hold #%i Elapsed time: %ss' %(ind+1,secondsToStr(current_time-time_zero)))
sys.stdout.flush()
renderplot(yvalues[-1],'end',0) # render final hold
pl.clf()
pl.close() # close figure to clear up memory
sys.stdout.write(' DONE\n\n')
sys.stdout.flush()

# run statistics
#print_prof_data()
#clear_prof_data()
gc.collect()
sys.stdout.write('\nEnd time: %s\n' %(time.strftime('%I:%M:%S %p')))
sys.stdout.write('\rElapsed time: %ss' %(secondsToStr(time.time()-time_zero)))
       
sys.stdout.write('\nfin.')
sys.stdout.flush()