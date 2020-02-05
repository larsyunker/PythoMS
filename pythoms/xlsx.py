"""
Class for opening and handling excel files with commonly used data formats
"""
import sys
import openpyxl as op
from openpyxl.utils.exceptions import InvalidFileException


class XLSX(object):
    def __init__(self,
                 bookname,
                 verbose=False,
                 create=False,
                 ):
        """
        A class for interacting with *xlsx (Microsoft Excel) files. This class requires the openpyxl package to
        function.

        :param str bookname: The name of the *xlsx file to load. The file extension is optional and the input is not
        case sensitive.

        :param bool verbose: Chatty mode.
        :param bool create: Whether to create the specified workbook if it cannot be located.

        **Examples**

        >>> xlfile = XLSX('book1')
        Loading workbook "book1.xlsx" into memory DONE
        >>> xlfile
        XLSX('book1.xlsx')
        """
        self.verbose = verbose
        self.wb, self.bookname = self.loadwb(bookname, create)

    def __str__(self):
        return 'Loaded excel file "%s"' % self.bookname

    def __repr__(self):
        return "%s('%s')" % (self.__class__.__name__, self.bookname)

    def add_cell(self, sheet, cellname, value, save=True):
        """
        Adds a value to the specified cell.

        :param str sheet: sheet name
        :param str cellname: cell name
        :param value: the valuve to inset
        :param bool save: whether to save after writing (default True)
        """
        if sheet in self.wb.sheetnames:  # check if sheet already present
            cs = self.get_sheet(sheet)
            cs[cellname] = value
        else:  # if not present, create
            cs = self.wb.create_sheet()
            cs.title = sheet
            cs[cellname] = value
        if save is True:  # if saving is desired
            self.save()

    def add_line(self, sheet, *values, save=True):
        """
        Adds the specified values to the next available line in the specified sheet

        :param str sheet: sheet name
        :param values:  values to insert
        :param bool save: whether to save after writing (default True)
        """
        if sheet in self.wb.sheetnames:  # check if sheet already present
            cs = self.get_sheet(sheet)
            nlines = cs.max_row
            for ind, val in enumerate(values):  # for each of the specified values
                cs.cell(
                    row=nlines + 1,
                    column=ind + 1,
                ).value = val
        else:  # if not present, create
            cs = self.wb.create_sheet()
            cs.title = sheet
            for ind, heading in enumerate(values):  # fill first row with values
                cs.cell(
                    row=1,
                    column=ind + 1,
                ).value = heading
        if save is True:  # if saving is desired
            self.save()

    def add_column(self, sheet, *values, save=True):
        """
        Adds the specified values to the next available column in the specified sheet.

        :param str sheet: sheet name
        :param values:  values to insert
        :param bool save: whether to save after writing (default True)
        """
        if sheet in self.wb.sheetnames:  # check if sheet already present
            cs = self.get_sheet(sheet)
            ncols = cs.max_column
            for ind, val in enumerate(values):  # for each of the specified values
                cs.cell(
                    row=ind + 1,
                    column=ncols + 1,
                ).value = val
        else:  # if not present, create
            cs = self.wb.create_sheet()
            cs.title = sheet
            for ind, val in enumerate(values):  # fill first row with values
                cs.cell(
                    row=ind + 1,
                    column=1,
                ).value = val
        if save is True:  # if saving is desired
            self.save()

    def checkduplicatesheet(self, sheet):
        """
        checks for duplicate sheets in the workbook and creates a unique name

        :param sheet: desired sheet name
        :return: unique sheet name
        :rtype: str
        """
        i = 1
        while sheet + ' (%s)' % str(i) in self.wb.sheetnames:
            i += 1
        return sheet + ' (%s)' % str(i)

    def correctextension(self, bookname):
        """attempts to correct the extension of the supplied workbook name"""
        oops = {'.xls': 'x', '.xl': 'sx', '.x': 'lsx', '.': 'xlsx', '.xlsx': ''}  # incomplete extensions
        for key in oops:
            if bookname.endswith(key) is True:
                bookname += oops[key]
                return bookname
        return bookname + '.xlsx'  # absent extension

    def evaluate(self, string, sheet):
        """
        Attempts to interpret and evaluate the contents of a cell containing basic transformations

        **Parameters**

        string: *string*
            The contents of a cell

        sheet: *string*
            The current sheet

        **Returns**

        return item: *type*
            description

        """
        if string.startswith('=') is False:  # if not an equation
            raise ValueError('The string %s is not an evaluable string' % string)
        contents = string[1:]

        rd = {  # replacement dictionary
            '^': '**',  # power
        }

        cs = self.get_sheet(sheet)

        evalstring = ''
        address = ''
        for ind, val in enumerate(contents):
            if val.isalpha() or val.isdigit():  # if part of an address
                address += val
                continue
            if val in rd:  # if in replacement dictionary
                if len(address) > 0:
                    cell_contents = cs[address]
                    # !! check if cell contents are a value, if not, call another instance of evaluate

                evalstring += rd[val]
            else:
                evalstring += val

        return eval(evalstring)

    def inds_to_cellname(self, row, col, lock=None):
        """
        Takes a pythonic index of row and column and returns the corresponding excel
        cell name.

        **Parameters**

        row: *integer*
            The pythonic index for the row, an index of 0 being the first row.

        col: *integer*
            The pythonic index for the column, an index of 0 being the first column.

        lock: 'cell','row', or 'col'
            If a locked cell is required, this can be specified here.
            e.g. If you wish to lock cell 'A1' to row, 'A$1' will be returned.
            Similarily, '$A1' is the column lock, and '$A$1' is the cell lock

        **Returns**

        cell name: *string*
            The excel-style cell name.


        **Examples**

        ::

            >>> XLSX.inds_to_cellname(55,14)
            'O56'
            >>> XLSX.inds_to_cellname(12,25)
            'Z13'
            >>> XLSX.inds_to_cellname(12,25,'cell')
            '$Z$13'
            >>> XLSX.inds_to_cellname(18268,558)
            'UM18269'


        **Notes**

        Based on http://stackoverflow.com/questions/23861680/convert-spreadsheet-number-to-column-letter
        """
        div = col + 1
        string = ""
        while div > 0:
            module = (div - 1) % 26
            string += chr(65 + module)
            div = int((div - module) / 26)
        if lock is not None:  # if the cell is to be locked
            out = ''
            if lock == 'col' or lock == 'cell':
                out += '$'
            out += string[::-1]  # string order must be reversed to be accurate
            if lock == 'row' or lock == 'cell':
                out += '$'
            out += str(row + 1)
            return out
        return string[::-1] + str(row + 1)

    def get_sheet(self, sheetname):
        """tries to retrieve the specified sheet name, otherwise returns None"""
        try:
            return self.wb[sheetname]
        except KeyError:
            return None

    def loadwb(self, bookname, create=False):
        """loads specified workbook into class"""
        if self.verbose is True:
            sys.stdout.write('\rLoading workbook "%s" into memory' % bookname)
        # cast path-like to string to enable extension check
        if type(bookname) is not str:
            bookname = str(bookname)
        try:
            wb = op.load_workbook(bookname)  # try loading specified excel workbook
        except (IOError, InvalidFileException):
            bookname = self.correctextension(bookname)  # attempts to correct the extension of the provided workbook
            if self.verbose is True:
                sys.stdout.write('\rLoading workbook "%s" into memory' % bookname)
            try:
                wb = op.load_workbook(bookname)
            except IOError:
                if self.verbose is True:
                    sys.stdout.write(' FAIL\n')
                if create is True:
                    """
                    Due to write-only mode, creating and using that book breaks the cell calls
                    The workbook is therefore created, saved, and reloaded
                    The remove sheet call is to remove the default sheet
                    """
                    if self.verbose is True:
                        sys.stdout.write('Creating workbook "%s" and loading it into memory' % bookname)
                    # wb = op.Workbook(bookname,write_only=False) # create workbook
                    wb = op.Workbook(bookname)  # create workbook
                    wb.save(bookname)  # save it
                    wb = op.load_workbook(bookname)  # load it
                    wb.remove(wb.worksheets[0])  # remove the old "Sheet"
                else:
                    raise IOError(
                        '\nThe excel file "%s" could not be found in the current working directory.' % (bookname))
        if self.verbose is True:
            sys.stdout.write(' DONE\n')
        return wb, bookname

    def pullmultispectrum(self, sheetname):
        """reads multispectrum output back into dictionary format"""
        cs = self.wb[sheetname]
        out = {}
        loc = 1
        while loc < cs.max_column:
            out[cs.cell(row=1, column=loc).value] = {
                'xunit': cs.cell(row=1, column=loc + 1).value,
                'yunit': cs.cell(row=1, column=loc + 2).value,
                'x': [],
                'y': [],
            }
            ind = 2
            while cs.cell(row=ind, column=loc + 1).value is not None:
                out[cs.cell(row=1, column=loc).value]['x'].append(cs.cell(row=ind, column=loc + 1).value)
                out[cs.cell(row=1, column=loc).value]['y'].append(cs.cell(row=ind, column=loc + 2).value)
                ind += 1
            loc += 4
        return out

    def pullrsim(self, sheet, TIC=True):
        """
        Pulls reconstructed single ion monitoring data from a specified sheet.

        **Parameters**

        sheet: *string*
            The excel sheet to pull the data from.

        TIC: *bool*
            Whether or not the second column contains total ion current data.
            See returns for more information.


        **Returns**

        data: *dictionary*
            A dictionary of dictionaries with each key corresponding to the
            name of the column (the value in row #1 of each column).
            Each subdictionary will be of the form
            ``data[key]['raw'] = [values]``
            and the number of values in the 'raw' subkey will be equal to
            the number in *time* and *TIC* returns.

        time: *list*
            A list of values in the first column of the sheet. The first
            column is assumed to be the x ordinate.

        TIC: *list*
            A list of values in the second column of the sheet. The second
            column is assumed to be the total ion current.
            **This will only be assumed and returned if the TIC keyword
            argument is True.**

        **Sheet data layout**

        This method assumes that the data contained in the sheet is arranged
        in the following fashion:

        ::

            Col A       Col B       Col C       Col D       ...
            x ordinate  TIC         name 1      name 2      ...
            x val 1     TIC value   1 value     2 value     ...
            ...         ...         ...         ...         ...


        Calling ``XLSX.pullrsim(sheetname,TRUE)`` would return the following:

        ::

            (
            {'name 1': {'raw': [value, ...]}, 'name 2': {'raw': [value, ...]}, ...},
            [x val 1, ...],
            [TIC value, ...]
            )

        Calling ``XLSX.pullrsim(sheetname,FALSE)`` would return the following:

        ::

            (
            {'TIC': {'raw': [TIC value, ...]}, 'name 1': {'raw': [value, ...]}, 'name 2': {'raw': [value, ...]}, ...},
            [x val 1, ...],
            )

        **See Also**

        This method is primarily used by PyRSIR.py. See this script for more details.

        """
        cs = self.wb[sheet]
        if TIC is True:
            tic = []
        time = []
        data = {}
        for col, colval in enumerate(cs.columns):
            for row, rowval in enumerate(colval):
                if row == 0:  # skip first row
                    continue
                elif colval[0].value == 'Time':  # if column is Time, append to that list
                    time.append(cs.cell(row=(row + 1), column=(col + 1)).value)
                elif colval[0].value == 'TIC' and TIC is True:  # if column is tic, append to that list
                    tic.append(cs.cell(row=(row + 1), column=(col + 1)).value)
                else:  # all other columns
                    if colval[row].value is not None:
                        if str(colval[0].value) not in data:
                            data['%s' % str(colval[0].value)] = {'raw': []}
                        data['%s' % str(colval[0].value)]['raw'].append(cs.cell(row=(row + 1), column=(col + 1)).value)
        if TIC is True:
            return data, time, tic
        else:
            return data, time

    def pullspectrum(
            self,
            sheet='spectrum',
            skiplines=0,
            column_offset=0,
    ):
        """
        Pulls an x,y set of paired values from the specified sheet.

        :param str sheet: The sheet name from which to extract the data.
        :param skiplines: If there are lines of data (e.g. acquisition parameters) before the name of the x and y data
            sets.
            e.g. a value of 0 means that the x and y column headers are
            in row 1, 2 means the headers are in row 2, etc.
        :param column_offset: If the data are not in columns 1 and 2, specify offset here
        :return: spectrum, x unit, y unit
        :rtype: ([[],[]], str, str)


        **Sheet data layout**

        This method assumes that the data contained in the sheet is arranged
        in the following fashion:

        ::

            Col A       Col B
            x unit      y unit
            x value     y value
            ...         ...


        Calling ``XLSX.pullspectrum('sheetname',0)`` on this data would return:

        ::

            (
            [[xvalue, ...], [yvalue, ...]],
            'x unit',
            'y unit'
            )

        **Examples**

        ::

            code line 1
            code line 2


        **See Also**

        optional

        """

        def tofloat(value, row, col):
            """attempts to convert to float and raises exception if an error is encountered"""
            try:
                return float(value)
            except ValueError:
                raise ValueError(
                    'The value "%s" (cell %s) in "%s" could not be interpreted as a float.\nCheck the value in this cell or change the number of lines skipped' % (
                        value, self.inds_to_cellname(row, col), self.bookname))

        if self.verbose is True:
            sys.stdout.write('Pulling spectrum from sheet "%s"' % sheet)
        skiplines -= 1
        specsheet = self.wb[sheet]
        spectrum = [[], []]
        for ind, row in enumerate(
                specsheet.rows):  # for each row append the mz and int values to their respective lists
            if ind > skiplines:  # skip specified number of lines
                if ind == skiplines + 1:  # header row
                    xunit = row[0 + column_offset].value
                    yunit = row[1 + column_offset].value
                    continue
                if row[0 + column_offset].value is not None and row[1 + column_offset].value is not None:
                    spectrum[0].append(tofloat(row[0 + column_offset].value, ind, 0))  # append values
                    spectrum[1].append(tofloat(row[1 + column_offset].value, ind, 1))
        if self.verbose is True:
            sys.stdout.write(' DONE\n')
        return spectrum, xunit, yunit

    def pullrsimparams(self, sheet='parameters'):
        """
        Loads reconstructed single ion monitoring parameters from the specified sheet.
        These are used in the PyRSIR.py script.

        **Parameters**

        sheet: *string*
            The sheet name that contains the data.


        **Returns**

        parameters: *dictionary*
            A dictionary containing the processing parameters for a PyRSIR excecution.


        **Valid column headings**

        There are several valid column headings that are recognized by this function.

        Name: name of the species
            This will be the dictionary key in the returned parameters dictionary.

        Formula: chemical formula
            The molecular formula for a given species can be provided, and the
            integration bounds will be automatically generated using the simulated
            isotope pattern. See ``Molecule.bounds()`` method for more details.

        Function: function in the mzML
            This can be specified if the function in the mass spec acquisition file
            is known. The function number can usually be viewed using the instrument
            manufacturers software or a program like ProteoWizard's seeMS.
            This is optional unless there is more than one spectrum type which matches
            the provided affinity.

        Affinity: the spectrum type where the species can be found
            This must be specified to indicate which mass spectrum type to look for
            the assigned species.
            Options: '+' (positive mode),'-' (negative mode), or 'UV' (UV-Vis channel).
            e.g. if the species is positively charged, specify '+'.
            If no affinity is specified, positive mode is assumed.

        Start: the integration start point
            The integration start point must be manually assigned if a molecular
            formula was not specified. The PyRSIR scipt will integrate values
            between the x values specified by *start* and *end*.
            If you wish to integrate only a single x value, specify the start
            value and leave the end value blank.

        End: the integration stop point
            The integration start point must be manually specified if a molecular
            formula was not provided.

        **Requirements for PyRSIR**

        At least one of *name*, *formula*, *start*, or *end* must be specified for
        PyRSIR to function.

        The excel sheet is expected to have the first row defining the columns
        (see above) and any subsequent row defining one species to track per
        line.

        """

        def othernames(oldsheet):
            """tries to find another common name for the parameters sheet"""
            others = ['Parameters', 'params', 'Params']
            for name in others:
                if name in self.wb.sheetnames:
                    return name
            raise KeyError('There is no "%s" sheet in "%s".' % (oldsheet, self.bookname))

        def lowercase(string):
            """tries to return the lowercase of the supplied value, and avoids NoneType errors of .lower()"""
            if string is None:
                return None
            return string.lower()

        if sheet not in self.wb.sheetnames:  # if the sheet can't be found, try other common names
            sheet = othernames(sheet)

        s = self.wb[sheet]  # load sheet in specified excel file

        names = ['name']  # valid name column headers
        formulas = ['formula', 'form', 'mf']  # valid molecular formula column headers
        affinities = ['affin', 'affinity']  # valid affinity column headers
        functions = ['fn', 'func', 'function']  # valid function column headers
        levels = ['level', 'lvl']  # valid level column headers
        smz = ['startmz', 'start mz', 'start m/z', 'startm/z', 'start']  # valid start m/z column headers
        emz = ['endmz', 'end mz', 'end m/z', 'endm/z', 'end']  # valid end m/z column headers

        self.rsimh = {}
        for ind, col in enumerate(s.columns):
            if lowercase(col[0].value) in names:
                self.rsimh['name'] = ind
            if lowercase(col[0].value) in formulas:
                self.rsimh['formula'] = ind
            if lowercase(col[0].value) in affinities:
                self.rsimh['affin'] = ind
            if lowercase(col[0].value) in functions:
                self.rsimh['function'] = ind
            if lowercase(col[0].value) in levels:
                self.rsimh['level'] = ind
            if lowercase(col[0].value) in smz:
                if 'bounds' not in self.rsimh:
                    self.rsimh['bounds'] = [None, None]
                self.rsimh['bounds'][0] = ind
            if lowercase(col[0].value) in emz:
                if 'bounds' not in self.rsimh:
                    self.rsimh['bounds'] = [None, None]
                self.rsimh['bounds'][1] = ind

        out = {}  # output dictionary
        for ind, row in enumerate(s.rows):
            if ind == 0:  # skip header row
                continue

            # figure out what to name the species
            name = None
            if 'name' in self.rsimh and row[self.rsimh['name']].value is not None:
                name = row[self.rsimh['name']].value  # if there is a name column and the value is not None
            elif 'formula' in self.rsimh and row[self.rsimh['formula']].value is not None:
                name = row[self.rsimh['formula']].value  # if there is instead a formula
            elif 'bounds' in self.rsimh and row[self.rsimh['bounds'][0]].value is not None:  # if there are bounds
                if row[self.rsimh['bounds'][1]].value is None:
                    name = '%.1f' % row[self.rsimh['bounds'][0]].value
                else:
                    name = '%.1f - %.1f' % (row[self.rsimh['bounds'][0]].value, row[self.rsimh['bounds'][1]].value)
            else:
                raise ValueError(
                    'A name could not be determined for row #%d\n%s' % (ind + 1, self.pullrsimparams.__doc__))
            out[name] = {}  # create name key

            for key in self.rsimh:  # go through the headers for that row and pull and values provided
                if key == 'name':
                    continue
                if key == 'bounds':
                    if row[self.rsimh[key][0]].value is not None:
                        out[name]['bounds'] = [float(row[self.rsimh[key][0]].value), None]
                    if row[self.rsimh[key][1]].value is not None:
                        out[name]['bounds'][1] = float(row[self.rsimh[key][1]].value)
                    continue
                if row[self.rsimh[key]].value is not None:
                    out[name][key] = row[self.rsimh[key]].value

        pos = ['+', 'pos', 'positive', 'Pos', 'Positive']  # positive mode valid inputs
        neg = ['-', 'neg', 'negative', 'Neg', 'Negative']  # negative mode valid inputs
        uv = ['UV', 'UV-Vis', 'UVVis', 'uv', 'uvvis', 'uv-vis']  # UV-Vis valid inputs
        for name in out:
            if 'affin' in out[name]:  # change affinity to valid ones used by mzML
                if out[name]['affin'] in pos:
                    out[name]['affin'] = '+'
                elif out[name]['affin'] in neg:
                    out[name]['affin'] = '-'
                elif out[name]['affin'] in uv:
                    out[name]['affin'] = '+'
                else:
                    raise ValueError('The affinity "%s" for species "%s" is not valid\n%s' % (
                        out[name]['affin'], name, self.pullrsimparams.__doc__))

            if 'function' in out[name]:  # if function is specified, convert to integer
                out[name]['function'] = int(out[name]['function'])

            if 'affin' not in out[name] and 'function' not in out[
                name]:  # if no affinity or function, assume 1st function
                out[name]['function'] = 1

            if 'level' in out[name]:  # if level is specified, convert to integer
                out[name]['level'] = int(out[name]['level'])

            if 'bounds' in out[name]:  # if bounds are specified, check that end is not less than start
                if out[name]['bounds'][1] is None:
                    pass
                elif out[name]['bounds'][0] > out[name]['bounds'][1]:
                    raise ValueError(
                        '\nThe end value is larger than the start value for species "%s".\nStart: %s\nEnd: %s\n%s' % (
                            name, str(out[name]['bounds'][0]), str(out[name]['bounds'][1]),
                            self.pullrsimparams.__doc__))

            if 'formula' not in out[name] and 'bounds' not in out[name]:
                raise ValueError('The species "%s" does not have a formula or integration bounds' % (name))

            if 'formula' not in out[name]:  # create formula key if not specified (required for pyrsir)
                out[name]['formula'] = None
        return out

    def removesheets(self, delete):
        """
        Removes a sheet from the excel workbook.

        **Parameters**

        delete: *string* or *list*
            The name(s) of the sheet to be deleted from the excel workbook.
            If a string is supplied, that sheetname will be deleted.
            If a list of strings is supplied, each sheetname in the list will
            be deleted.


        **Returns**

        return item: *type*
            description


        **Examples**

        ::

            code line 1
            code line 2


        **See Also**

        optional

        """
        """
        removes sheets from the excel file
        delete is a set of strings to be removed
        """
        if type(delete) is str:  # if a single string is provided
            delete = [delete]
        for sheet in self.wb.sheetnames:  # clears sheets that will contain new peak information
            if sheet in delete:
                dels = self.wb[sheet]
                self.wb.remove_sheet(dels)

    def save(self, outname=None):
        """
        Commits changes to the workbook.

        **Parameters**

        outname: *string*
            Allows specification of a separate workbook to save as.
            If this is left as None (default), the provided filename
            given on initialization will be used.


        """
        def version_input(string):
            """checks the python version and uses the appropriate version of user input"""
            import sys
            if sys.version.startswith('3.'):
                return input('%s' % string)
            else:
                raise EnvironmentError('The version_input method encountered an unsupported version of python.')

        if outname is None:
            outname = self.bookname

        try:
            self.wb.save(outname)
        except IOError:
            version_input(
                '\nThe excel file could not be written. Please close "%s" and press any key to retry save.' % outname)
            try:
                self.wb.save(outname)
            except IOError:
                raise IOError('\nThe excel file "%s" could not be written.' % outname)

    def updatersimparams(self, sp, sheet='parameters'):
        """
        Updates reconstructed single ion monitoring parameters.
        This is usually run at the end of the PyRSIR script in order to
        save additional processing parameters that the script determined.

        **Parameters**

        sp: *dictionary*
            Dictionary of species data after PyRSIR processing.

        sheet: *string*
            Allows specification of a specific sheet to save as.

        """
        if sheet not in self.wb.sheetnames:  # if the sheet can't be found, try other common names
            sheet = self.pullrsimparams.othernames(sheet)

        s = self.wb[sheet]  # load sheet in specified excel file

        for ind, row in enumerate(s.rows):
            if ind == 0:
                continue
            # determine name key to look for
            if row[self.rsimh['name']].value is not None:  # if name is set
                key = row[self.rsimh['name']].value
            elif row[self.rsimh['formula']].value is not None:  # if formula is set
                key = row[self.rsimh['formula']].value
            elif row[self.rsimh['bounds'][0]].value is not None:
                if row[self.rsimh['bounds'][1]].value is None:
                    key = '%.1f' % row[self.rsimh['bounds'][0]].value
                else:
                    key = '%.1f - %.1f' % (row[self.rsimh['bounds'][0]].value, row[self.rsimh['bounds'][1]].value)
            else:
                key = str(row[0].value)
            for hkey in self.rsimh:  # for the defined column headers in the rsim headers dictionary
                if hkey in sp[key]:  # if the species has that key
                    if hkey == 'bounds':
                        start, end = self.rsimh[hkey]
                        row[start].value = sp[key][hkey][0]
                        row[end].value = sp[key][hkey][1]
                    elif row[self.rsimh[
                        hkey]].value is None:  # if the cell has not been filled (avoids overwriting user specifications)
                        row[self.rsimh[hkey]].value = sp[key][hkey]
                        # if row[2].value is None: # if no affinity defined
                        #    row[2].value = sp[key]['affin']
                        # if row[3].value is None: # if there are no bounds, update bounds
                        #    row[3].value = sp[key]['bounds'][0]
                        # if row[4].value is None:
                        #    row[4].value = sp[key]['bounds'][1]
                        # if sp[key]['match'] is not None:
                        #    if s.cell(row=1,column=6).value is None:
                        #        s.cell(row=1,column=6).value = 'std err of reg'
                        #    s.cell(row=ind+1,column=6).value = sp[key]['match']

    def writemultispectrum(self, xlist, ylist, specname, xunit='x', yunit='y', sheetname='spectra', chart=False):
        """
        Writes multiple spectra to a single workbook sheet.
        This can be called multiple times, and the class instance will remember
        the current location of the last inserted spectrum, and automatically
        determine the next columns to write into.

        **Parameters**

        xlist: *list*
            List of x values.

        ylist: *list*
            List of y values. This is assumed to have the same dimensions as x
            as well as be paired with x.

        xunit: *string*
            The unit of the x values. This will be placed at the top of the
            x column for the provided spectrum.

        yunit: *string*
            The unit of the y values. This will be placed at the top of the
            y column for the provided spectrum.

        sheetname: *string*
            The name of the sheet to write data to.

        specname: *string*
            The name of the spectrum being written. This will be inserted beside
            the spectrum for ease of identification.

        chart: *bool*
            Whether or not to output a chart of the data.


        **Output data layout**

        specname|xunit   |yunit   | blank |...repeated
        blank   |xvalues |yvalues | blank |...repeated
        ...     |...     |...     | blank |...repeated

        """
        if 'wms' not in self.__dict__:  # check for dictionary in self
            self.wms = {sheetname: 1}
        if sheetname not in self.wms:  # check for preexisting key
            self.wms[sheetname] = 1

        if sheetname not in self.wb.sheetnames:  # if sheet does not exist
            cs = self.wb.create_sheet()
            cs.title = sheetname
        else:
            cs = self.wb[sheetname]  # load existing sheet

        cs.cell(row=1, column=self.wms[sheetname]).value = specname
        cs.cell(row=1, column=self.wms[sheetname] + 1).value = xunit
        cs.cell(row=1, column=self.wms[sheetname] + 2).value = yunit
        for ind, val in enumerate(xlist):
            cs.cell(row=ind + 2, column=self.wms[sheetname] + 1).value = val  # write x value
            cs.cell(row=ind + 2, column=self.wms[sheetname] + 2).value = ylist[ind]  # write y value

        if chart is True:
            chart = op.chart.ScatterChart()  # generate the chart object
            chart.title = str(specname)  # convert title to string to avoid TypeError
            xvals = op.chart.Reference(cs, min_col=self.wms[sheetname] + 1, min_row=2,
                                            max_row=len(xlist) + 1)  # define the x values
            chart.x_axis.title = xunit  # x axis title
            yvals = op.chart.Reference(cs, min_col=self.wms[sheetname] + 2, min_row=2, max_row=len(xlist) + 1)
            chart.y_axis.title = yunit  # y axis title
            series = op.chart.Series(yvals, xvals)
            chart.series.append(series)  # add the data to the chart
            if self.wms[sheetname] // 4 % 2 == 0:  # alternate the location of the output charts
                shifty = 1
            else:
                shifty = 16
            cs.add_chart(chart,
                         self.inds_to_cellname(shifty, self.wms[sheetname] - 1))  # add the chart to the worksheet

        self.wms[sheetname] += 4  # shift location over 4

    def writersim(self, sp, time, key, sheetname, mode, tic=None):
        """
        Writes Reconstructed Single Ion Monitoring data from the PyRSIR script
        to a sheet in the workbook. (See PyRSIR.py for more details)

        **Parameters**

        sp: *dictionary*
            The dictionary of processed values resulting from PyRSIR

        time: *list*
            List of time values corresponding to the scans.

        key: *string*
            This is the name of the subkey within a given species'
            dictionary containing the list to be written.
            e.g. for raw data this would be 'raw' and for 3-summed
            data this would be '3sum'.
            This is part of determing the full PyRSIR dictionary
            key of the data.

        sheetname: *string*
            The name of the sheet to write the data to.

        mode: *string*
            The current mode being output (either '+','-', or 'UV').
            These allow differentiation of the various acquisition
            modes that may have been stored in a given mass spec
            run.
            This is part of determining the full PyRSIR dictionary
            key of the data.

        tic: None or *list*
            If provided, this is a list of total ion current values
            usually used for normalization of mass spectrometric data.


        **Notes**

        If the specified sheet name is already present in the workbook,
        the data will not be written.

        """
        if sheetname not in self.wb.sheetnames:
            cs = self.wb.create_sheet()  # create new sheet
            cs.title = sheetname  # rename sheet
            cs['A1'] = 'Time'
            offset = 0
            if tic is not None:
                offset += 1  # skip TIC column
                cs['B1'] = 'TIC'
            for ind, val in enumerate(time):  # write time information
                cs.cell(row=(ind + 2), column=1).value = time[ind]  # write time list
                if tic is not None:
                    cs.cell(row=(ind + 2), column=2).value = tic[ind]  # write TIC list
            col = 1 + offset
            for species in sorted(sp.keys(), key=lambda x: str(x)):
                dct = sp[species]
                if dct['affin'] is mode:  # if the species' affinity is the mode
                    col += 1  # +1 from 0 to 1, +1 each to skip Time and TIC columns
                    cs.cell(row=1, column=col).value = str(species)  # write species names
                    for ind, val in enumerate(dct[key]):
                        cs.cell(row=(ind + 2), column=col).value = val

    def writespectrum(self, x, y, sheet='spectrum', xunit='m/z', yunit='counts', norm=True, chart=True):
        """
        Writes an x,y spectrum to the specified sheet in the workbook.

        **Parameters**

        x: *list*
            List of x values.

        y: *list*
            List of y values, paired with x and of the same length as the x list.

        sheet: *string*
            The name of the sheet to write the data to in the workbook.

        xunit: *string*
            The units of the x values. This will be inserted at the top
            of the x column (Column A).

        y unit: *string*
            The units of the y values. This will be inserted at the top
            of the y column (Column B).

        norm: *bool*
            Whether or not to output a third column of normalized y values.
            The values will be normalized to 1.

        chart: *bool*
            Whether or not to plot the spectrum data as a chart.

        """
        if sheet in self.wb.sheetnames:
            sheet = self.checkduplicatesheet(sheet)
        ws = self.wb.create_sheet()
        ws.title = sheet
        ws['A1'] = xunit
        ws['B1'] = yunit
        if norm is True:
            ws['C1'] = 'Normalized'
            ws['D1'] = 'Maximum y value'
            # ws['D2'] = max(y)
            ws['D2'] = '=MAX(B:B)'
        for ind, val in enumerate(x):
            ws[self.inds_to_cellname(1 + ind, 0)] = x[ind]
            ws[self.inds_to_cellname(1 + ind, 0)].number_format = '0.000'
            ws[self.inds_to_cellname(1 + ind, 1)] = y[ind]
            ws[self.inds_to_cellname(1 + ind, 1)].number_format = '0'
            if norm is True:
                ws[self.inds_to_cellname(1 + ind, 2)] = '=%s/$D$2' % self.inds_to_cellname(1 + ind, 1)
        if chart is True:  # if a chart object is called for
            chart = op.chart.ScatterChart()  # generate the chart object
            xvals = op.chart.Reference(ws, min_col=1, min_row=2, max_row=len(x) + 1)  # define the x values
            chart.x_axis.title = 'm/z'  # x axis title
            if norm is False:  # if there is no normalized data
                yvals = op.chart.Reference(ws, min_col=2, min_row=1, max_row=len(x) + 1)
                chart.y_axis.title = 'Intensity (counts)'
            if norm is True:  # if there is normalized data
                yvals = op.chart.Reference(ws, min_col=3, min_row=1, max_row=len(x) + 1)
                chart.y_axis.title = 'Normalized Intensity'
            series = op.chart.Series(yvals, xvals)
            chart.series.append(series)  # add the data to the chart
            ws.add_chart(chart, 'E1')  # add the chart to the worksheet


if __name__ == '__main__':
    name = 'Useless delete this'
    xlfile = XLSX(name, create=True)

